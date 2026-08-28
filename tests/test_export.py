import io
from datetime import datetime

import openpyxl
import pandas as pd

from core.export import build_excel, rows_to_frame
from core.models import CartItem, OrderRow
from core.settings import KST

COLUMNS_MAP = {
    "주문번호": "order_number",
    "상품명": "product_name",
    "옵션명": "option_name",
    "수량": "quantity",
    "단가": "unit_price",
    "주문금액": "amount",
    "수령자": "customer_name",
    "주소": "full_address",
    "우편번호": "zip_code",
}


def _order(order_id="o1", items=None, zip_code=None, address_detail=None):
    return OrderRow(
        id=order_id,
        broadcast_id="b1",
        order_number="20260830001",
        staff_name="민지",
        customer_name="홍길동",
        phone="010-1234-5678",
        address="서울 강남구 테헤란로 1",
        address_detail=address_detail,
        full_address="서울 강남구 테헤란로 1" + (f" {address_detail}" if address_detail else ""),
        zip_code=zip_code,
        created_at=datetime(2026, 8, 30, 20, 0, tzinfo=KST),
        created_at_kst="2026-08-30 20:00:00",
        items=items or [],
    )


def _items():
    return [
        CartItem(product_name="가디건", option_name="그레이", unit_price=78000, quantity=2),
        CartItem(product_name="스커트", option_name="단일", unit_price=129000, quantity=1),
    ]


def test_rows_to_frame_item_mode_expands_one_row_per_item():
    frame = rows_to_frame([_order(items=_items())], COLUMNS_MAP, row_mode="item")
    assert len(frame) == 2
    assert list(frame["상품명"]) == ["가디건", "스커트"]
    assert list(frame["수량"]) == [2, 1]
    assert list(frame["주문번호"]) == ["20260830001", "20260830001"]


def test_rows_to_frame_item_mode_computes_order_amount():
    frame = rows_to_frame([_order(items=_items())], COLUMNS_MAP, row_mode="item")
    # 주문금액 = 단가 × 수량
    assert list(frame["주문금액"]) == [78000 * 2, 129000 * 1]


def test_rows_to_frame_order_mode_collapses_items_into_one_row():
    frame = rows_to_frame([_order(items=_items())], COLUMNS_MAP, row_mode="order")
    assert len(frame) == 1
    assert frame.iloc[0]["상품명"] == "가디건 그레이 x2; 스커트 단일 x1"
    assert frame.iloc[0]["수량"] == 3
    assert frame.iloc[0]["단가"] is None
    # 주문 1행 모드의 주문금액은 전체 합계
    assert frame.iloc[0]["주문금액"] == 78000 * 2 + 129000 * 1


def test_rows_to_frame_uses_full_address():
    frame = rows_to_frame([_order(items=_items(), address_detail="101동 202호")], COLUMNS_MAP)
    assert frame.iloc[0]["주소"] == "서울 강남구 테헤란로 1 101동 202호"


def test_rows_to_frame_handles_order_without_items():
    frame = rows_to_frame([_order(items=[])], COLUMNS_MAP, row_mode="item")
    assert len(frame) == 1
    assert frame.iloc[0]["상품명"] is None


def test_rows_to_frame_null_field_produces_empty_column():
    columns_map = {**COLUMNS_MAP, "메모": None}
    frame = rows_to_frame([_order(items=_items())], columns_map)
    assert frame["메모"].isna().all()


def test_build_excel_zip_column_is_text_formatted():
    frame = rows_to_frame([_order(items=_items(), zip_code="06134")], COLUMNS_MAP)
    excel_bytes = build_excel(frame, None, "주문내역", "검토필요")

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["주문내역"]
    zip_col_idx = frame.columns.get_loc("우편번호") + 1
    for row in ws.iter_rows(min_row=2, min_col=zip_col_idx, max_col=zip_col_idx):
        assert row[0].number_format == "@"


def test_build_excel_includes_review_sheet_when_non_empty():
    frame = rows_to_frame([_order(items=_items())], COLUMNS_MAP)
    review_frame = pd.DataFrame({"주문번호": ["20260830002"], "주소": ["미상"]})

    excel_bytes = build_excel(frame, review_frame, "주문내역", "검토필요")

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "검토필요" in wb.sheetnames


def test_build_excel_omits_review_sheet_when_empty():
    frame = rows_to_frame([_order(items=_items())], COLUMNS_MAP)
    review_frame = pd.DataFrame({"주문번호": []})

    excel_bytes = build_excel(frame, review_frame, "주문내역", "검토필요")

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "검토필요" not in wb.sheetnames
